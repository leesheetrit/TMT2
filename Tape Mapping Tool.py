
#####import libraries
import tkinter as tk
import openpyxl as xl
from tkinter import ttk
import tkinter.font as tkfont
import pandas as pd
from tkinter import *
from sqlalchemy import create_engine


global qrm_db  #create engine
qrm_db = create_engine('mssql+pyodbc://CFAWSQRMSQL01/Servicing?driver=SQL+Server+Native+Client+11.0',
    fast_executemany = True)

def btn_temp():
    print('pass')

#this works as long as the drop down list isn't displayed
def keyboard_first_letter_select(event):
    combo = event.widget

    # Get the typed letter from the event
    letter = event.char.lower()

    # Find the first list item that starts with the typed letter
    for i, item in enumerate(combo['values']):
        if item.lower().startswith(letter):
            combo.current(i)  # Select the found item
            break

def read_TMT():
    
    ####Load Tape Mapping Tool Excel File
    print('Loading Excel File...')
    global wb_path
    wb_path = 'I:/ServicingMSR/QRM/Trades Upload Files/Tape Mapping Tool/Tape Mapping Tool.xlsx'
    global wb
    wb = xl.load_workbook(wb_path)
    global dest_columns_sheet
    dest_columns_sheet = wb['Destination Columns']
    global tape_sheet
    tape_sheet = wb['Tape']
    global dest_field_values_sheet
    dest_field_values_sheet = wb['Destination Field Values']
    global values_mapping_sheet
    values_mapping_sheet = wb['ValuesMapping']

    #Find the last used row on the Excel sheet
    print('Finding Last Row on Tape File...')
    global last_row
    last_row = tape_sheet.max_row
    
    ####Read Source Columns 
    # Create a dictionary to store the source columns
    #stores the source column letter as the Dictionary item and the source Column Name as the dictionary value

    global source_columns
    source_columns = {}
    for cell in tape_sheet[1]:
        if cell.value is not None:
            source_columns[cell.column_letter] = cell.value
                
    print('Source Columns Read.')          
    # Sort the source column names alphabetically
    #source_column_names = sorted(source_columns.values())

    global source_column_names
    source_column_names = sorted(map(str, source_columns.values()))

    print('Source Columns Sorted.')          

    ####Read Destination Columns 
    #Read the destination columns into a dictionary called "destination_columns"
    #dictionarys are Key-Value pairs
    #stores the destination column name as the Dictionary Key
    global destination_columns
    destination_columns = {}
    for row in dest_columns_sheet.iter_rows(min_row=2, min_col=0, max_col=4):
        dest_col_name = row[0].value
        requires_mapping = row[1].value
        dest_value_example = row[2].value
        is_percentage = row[3].value
        
        #the value for the destination_columns dictionary is another dictionary
        #the key-value pairs of the child dictioary as shown below
        if dest_col_name is not None:
            destination_columns[dest_col_name] = {
               'requires_mapping': requires_mapping,
               'dest_value_example': dest_value_example,
               'is_percentage': is_percentage,
               }

def write_field_mappings():
####Write Field mappings
    # Print the index and the values of the variables in DivideBy100Vars
    for idx, var in enumerate(DivideBy100Vars):
        # print(f"DivideBy100Vars: Index: {idx}, Value: {var}")
        if var is False:
            print(f"DivideBy100Vars: Index: {idx}, Value: {var}")
        else:
            print(f"DivideBy100Vars: Index: {idx}, Value: {var.get()}")      
    i=-1
    # Write the destination column names going across the columns of last_row
    for i, (item, var, requires_mapping,is_percentage) in enumerate(column_mapping_rules):
      tape_sheet.cell(row=last_row+2, column=i+1).value = item
    print('Destination Column Labels writted into XLS')
    
    i=0
    row = last_row + 3
    # Skip to the next row and write the associated source column formulas or mappings across that row, one per column
    print('All Loans are Current equals:', AllCurrentVar.get())
    for i, (item, var, requires_mapping,is_percentage) in enumerate(column_mapping_rules):
        col_letter = None
        for letter, name in source_columns.items():
            if name == var.get():
                col_letter = letter
                break
            
        print('Evaluating logic for',item,'All current: ',AllCurrentVar.get(), "i equals ", i)

        if DivideBy100Vars[i] is False:
            #no checkbox for this item
            print(f"Item: {item}, Index: {i}, DivideBy100Vars Checkbox Value: {DivideBy100Vars[i]}")
        else:
            #read whether the checkbox was checked or not
            print(f"Item: {item},  Index: {i},DivideBy100Vars Checkbox Value: {DivideBy100Vars[i].get()}")
             
        #if a source column is selected
        if col_letter is not None or (item =='Delq' and AllCurrentVar.get()):
            #if the checkbox for all curent is checked
            if item == "Delq" and AllCurrentVar.get():
                #write all current
                formula = "PREPAID OR CURRENT"
            elif requires_mapping== 'TRUE' and col_letter is not None:
                print('vlookup for ',item)
                formula = f'=VLOOKUP({col_letter}$1'+ ' & "-" & ' +f'{col_letter}2 & "-" & "{item}", ValuesMapping!$A:$B, 2, FALSE)'       
            #no divide by 100 Checkbox exists
            elif DivideBy100Vars[i] is False:
                formula = f"={col_letter}2"
            else:
                #if divide by 100 is checked
                if DivideBy100Vars[i].get():
                    formula = f"={col_letter}2/100"
                #divide by 100 is not checked
                else:
                    formula = f"={col_letter}2"
                
            tape_sheet.cell(row=row, column=i+1).value = formula
    
    # Save the changes to the Excel file
    wb.save(wb_path)
    print('Mappings Written')
 
def read_unique_values_from_excel(filename, sheet_name, column_name, max_row, max_values=15):
    #### functions for Values Mapping Window  read unique source column values 
    df = pd.read_excel(filename, sheet_name=sheet_name, nrows=max_row)
    column_values = df[column_name].unique()  # Get unique values from the specified column
    capped_values = column_values[:max_values]  # Cap the values at the specified maximum   
    
    return capped_values.tolist()  # Return the capped values as a list

def write_value_mappings():
    #### Write values Mappings
    i=0
    for i, (source_col_letter,source_col_name,unique_source_value1,dest_field_name,dest_field_value) in enumerate(value_mapping_rules):
     
        values_mapping_sheet.cell(row=i+2, column=1).value = str(source_col_name) + '-' + str(unique_source_value1) + '-' + str(dest_field_name)
        print('writing')
        print(dest_field_name)
        values_mapping_sheet.cell(row=i+2, column=2).value = dest_field_value.get()
        
    wb.save(wb_path)
    print('Values mappings writted into excel')

def save_field_mappings():
    #loop through each source column and destination column in the column_mapping_rules list
    #append to a SQL server table called "TMT_Field_Mappings". First column is Sellers Names, Second column is Mappings Name, third column is source column, furth coulmn is destination column 
   
    sellers_name = seller_entry.get()
    mappings_name = mapping_name_entry.get()
    
    
    #not quite working, write PY_VAR849 instead of source field name
    for i, (dest_field, source_field, requires_mapping,is_percentage) in enumerate(column_mapping_rules):
       
        if source_field.get() is not None and source_field.get() != "":
            insert_query = f"INSERT INTO Servicing.dbo.TMT_Field_Mappings \
                (Sellers_Name, Mappings_Name, Destination_Field, Source_Field) \
                    VALUES \
                        ('{sellers_name}', '{mappings_name}', '{dest_field}', '{source_field.get()}')"
            qrm_db.connect().execute(insert_query)
            print('Wrote: ', dest_field,' - ', source_field.get())

    print('Field Mappings Save Complete')
    
def open_save_load_field_mappings_window():
    #create a new tkinter toplevel
    sl_field_mappings_window = tk.Toplevel(field_map_window)
    sl_field_mappings_window.title('Save and Load Mapping Rules')
    sl_field_mappings_window.geometry('300x300+50+50')
    
    #col 1- Save Mappings
    label = tk.Label(sl_field_mappings_window, text='Save Mappings', font=bold_font)
    label.grid(row=0, column=0,padx=10)
    label = tk.Label(sl_field_mappings_window, text="Enter Broker or Seller's Name")
    label.grid(row=1, column=0)
    global seller_entry
    seller_entry = tk.Entry(sl_field_mappings_window)
    seller_entry.grid(row=2, column=0)
    seller_entry.insert(0, "seller1") 
    label = tk.Label(sl_field_mappings_window, text="Enter Mappings Name")
    label.grid(row=3, column=0)
    global mapping_name_entry
    mapping_name_entry = tk.Entry(sl_field_mappings_window)
    mapping_name_entry.grid(row=4, column=0)
    mapping_name_entry.insert(0, "newMapping1") 
    button = tk.Button(sl_field_mappings_window, text="Save Mapping Rules", command=save_field_mappings)  
    button.grid(row=5, column=0)
    
    
    #col 2- Load Mappings
    label = tk.Label(sl_field_mappings_window, text='Load Mappings', font=bold_font)
    label.grid(row=0, column=1)
    listbox = tk.Listbox(sl_field_mappings_window)
    listbox.grid(row=1, column=1, columnspan=2)
    
    
    #in work***
    # Fetch distinct combinations of Sellers_Name and Mappings_Names
    print('Reading Mappings from SQL-Server')
    query = "SELECT DISTINCT Sellers_Name, Mappings_Name FROM Servicing.dbo.TMT_Field_Mappings"
    result = qrm_db.execute(query)
    
    # Populate listbox with distinct combinations
    for row in result:
        #consider creating a list of some sort for easier accessibility later
        listbox.insert(tk.END, f"{row[0]} - {row[1]}")
    
    button = tk.Button(sl_field_mappings_window, text="Load Mapping Rules", command=btn_temp)    
    button.grid(row=2, column=1)

    #create a listbox 
    #the list box is to be populate by the distinct combinations of Sellers_Name and Mappings_Names which are the first two columns in the SQLServertable TMT_Field_Mappings
    #



def open_value_mappings_window():
    #### create the value mapping window
    value_mapping_window = tk.Toplevel(field_map_window)
    value_mapping_window.title('Value Mapping Rules')
    #width, height, 50 pixels from left edge, 50 pixeles down from top edge
    value_mapping_window.geometry('900x900+50+50')
   
    #code for vertical scroll bar 
    #https://github.com/flatplanet/Intro-To-TKinter-Youtube-Course/blob/master/full_scroll.py
    # Create A Main Frame
    print('Creating New Window...')
    main_frame = tk.Frame(value_mapping_window)
    main_frame.pack(fill=BOTH, expand=1)
   
    # Create A Canvas
    my_canvas = Canvas(main_frame)
    my_canvas.pack(side=LEFT, fill=BOTH, expand=1)
   
    # Add A Scrollbar To The Canvas
    my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
    my_scrollbar.pack(side=RIGHT, fill=Y)
   
    # Configure The Canvas
    my_canvas.configure(yscrollcommand=my_scrollbar.set)
    my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion = my_canvas.bbox("all")))
    
    # Create ANOTHER Frame INSIDE the Canvas
    second_frame = Frame(my_canvas)
    
    # Add that New frame To a Window In The Canvas
    my_canvas.create_window((0,0), window=second_frame, anchor="nw")
   
    # Loop through destination_columns and create labels for columns that requires value mapping
    i= 0
    row_num = 0 
    
    #set the max row for source field values
    #checks if the column mappings have already been written, then limits the max row to the end of the tape (ignors the new mappinss)
    print('Checking if Column Mappings have been Written...')
    cell_value = tape_sheet["A{}".format(last_row-1)].value
    if cell_value == 'LOAN NUMBER' :
          last_row_adjusted = last_row - 4
    else:
          last_row_adjusted = last_row
          
    print('Last Row for Source Values is ',last_row_adjusted)
    
    #iterate through the destination columns list
    for i, (dest_field_name, var, requires_mapping,is_percentage) in enumerate(column_mapping_rules):
     
     if requires_mapping == 'TRUE': 
         #create a label for each destination column
         dest_col_label = tk.Label(second_frame, text="Dest: " + dest_field_name)
         dest_col_label.grid(row=row_num, column=1)
        
         # List to store destination field values
         dest_field_values = []
        
         # Iterate through each row in the destination field values sheet
         print('Reading Destination Field Values for ', dest_field_name)
         for row in dest_field_values_sheet.iter_rows(min_row=2, values_only=True):  
             xl_dest_field_name = row[0]  # Column A
             dest_field_value1 = row[1]  # Column B
        
             if xl_dest_field_name == dest_field_name:
                 dest_field_values.append(dest_field_value1)
        
         # Print the list of matching destination field values
         print('Identified Destination values: ', dest_field_values)
        
        
         print('Finding Matching Source Column...')
        #find the matching source columns
         for source_col_letter, source_col_name in source_columns.items():
             if source_col_name == var.get():
                 print('Source Column Found')
                 #create a label for matchin source column
                 source_col_label = tk.Label(second_frame, text="Source: " + source_col_name)
                 source_col_label.grid(row=row_num, column=0)
                
                 #create a unique list of source values, capped at 15 values
                 ####call read_unique_values_from_excel
                 print('Reading Source Values for ', source_col_name)
                 unique_source_values = read_unique_values_from_excel(filename= wb_path, sheet_name = 'Tape', column_name = source_col_name, max_row = last_row_adjusted,  max_values = 15)
                 print('Identifiend Source Values: ',unique_source_values)
                
                 print('Creating Labels for ', source_col_name)
                 #create a label for each unique destination value
                 for unique_source_value1 in unique_source_values:
                     row_num += 1
                      # Create a label for each unique destination value
                     label = tk.Label(second_frame, text=unique_source_value1)
                     label.grid(row=row_num, column=0)
                    
                     var = tk.StringVar(value_mapping_window)
                     # Create a combo box for each unique destination value, populate it with the source values
                     combo = ttk.Combobox(second_frame, textvariable=var, values=list(dest_field_values), width=40, state ="readonly",height = 20)
                     combo.grid(row=row_num, column=1)
                    
                     for dest_field_value1 in dest_field_values:
                         #error handling for float
                        if isinstance(unique_source_value1, float) or isinstance(dest_field_value1, float):
                            if unique_source_value1 == dest_field_value1:
                                combo.set(dest_field_value1)
                                break
                        else:
                            #case insensitive match
                            if unique_source_value1.lower() == dest_field_value1.lower():
                                combo.set(dest_field_value1)
                                break

                      # #add destination column, the selected variable, and whether or not the column requires mapping to the "column_mapping_rules" list
                     value_mapping_rules.append((source_col_letter,source_col_name,unique_source_value1,dest_field_name,var))
                     
                 break

         row_num += 1
        
    #button
    button = tk.Button(second_frame, text="Write Value Mapping", command=write_value_mappings)
    button.grid(row=row_num, column=0, columnspan=2)
    


def open_field_map_window():
    #### Create Column Mapping Tk window 
    
    #read excel file
    read_TMT()
    
    global field_map_window
    field_map_window = tk.Toplevel(MainNavWindow)
    field_map_window.title('Column Mapping Rules')
    print('TKinter window created')
    row_num = 0 #variable for row in the tkinter GUI

    label = tk.Label(field_map_window, text='Destination Column', font=bold_font)
    label.grid(row=row_num, column=0)
    label2= tk.Label(field_map_window, text='Requires Mapping', font=bold_font)
    label2.grid(row=row_num, column=1)
    label3= tk.Label(field_map_window, text='One Loan Example', font=bold_font)
    label3.grid(row=row_num, column=2)
    label3= tk.Label(field_map_window, text='Source Column', font=bold_font)
    label3.grid(row=row_num, column=3)
    label4= tk.Label(field_map_window, text='Special', font=bold_font)
    label4.grid(row=row_num, column=4)
    
    #temporarily showing this at the top
    row_num = 1
    button = tk.Button(field_map_window, text="Save Mapping Rules", command=open_save_load_field_mappings_window)    
    # button.grid(row=row_num, column=0, columnspan=2)
    button.grid(row=row_num, column=1)
    
    row_num = 2
    #create a "combos" list
    global column_mapping_rules
    column_mapping_rules = []
    
    global value_mapping_rules
    value_mapping_rules = []
    
    #### Gen. Widgets for Col. Mapping Window
    # Create a list to store the DivideBy100Var variables
    
    global DivideBy100Vars
    DivideBy100Vars = []
    global AllCurrentVar
    AllCurrentVar = None 
    
    for item, child_dictionary in destination_columns.items():
        # Create a label and combo box for each Destination Column in the "destination_columns" 
        label = tk.Label(field_map_window, text=item)
        label.grid(row=row_num, column=0)
        
        if child_dictionary['requires_mapping'] == 'FALSE':
            RequiresMappingLabelText = ''
        else:
            RequiresMappingLabelText= child_dictionary['requires_mapping']
        #label2= tk.Label(field_map_window, text=child_dictionary['requires_mapping'])
        label2= tk.Label(field_map_window, text=RequiresMappingLabelText)
        label2.grid(row=row_num, column=1)
        
        label3= tk.Label(field_map_window, text=child_dictionary['dest_value_example'])
        label3.grid(row=row_num, column=2)
        
        #source columns combo box
        var = tk.StringVar(field_map_window)
        combo = ttk.Combobox(field_map_window, textvariable=var, values=list(source_column_names), width=40, state ="readonly",height = 20)
        #combo = ttk.Combobox(field_map_window, textvariable=var, values=list(source_column_names), width=40,height = 20)
        combo.grid(row=row_num, column=3)
        combo.bind('<KeyPress>', keyboard_first_letter_select)
           
        # Check if any of the source column names match "item" and set it as the default value
        #comment this out if you don't want the default matching
        for source_column_name in source_column_names:
            if source_column_name.lower() == item.lower():
                var.set(source_column_name)
                break
        
        if child_dictionary['is_percentage'] == True:
            # Create a variable to hold the state of the checkbox
            DivideBy100Var = tk.BooleanVar()
            # Create the checkbox
            checkbox = tk.Checkbutton(field_map_window, text='Divide by 100?', var=DivideBy100Var)
            checkbox.grid(row=row_num, column=4)
            DivideBy100Vars.append(DivideBy100Var) # Add the DivideBy100Var to the DivideBy100Vars list
            # DivideBy100Checkboxes.append(checkbox)
        elif item == 'Delq':
            AllCurrentVar = tk.BooleanVar()
            checkbox = tk.Checkbutton(field_map_window, text='All Current?', var=AllCurrentVar)
            checkbox.grid(row=row_num, column=4)
            DivideBy100Vars.append(False)
            # DivideBy100Checkboxes.append(False)
        else:
            DivideBy100Vars.append(False)
            # DivideBy100Checkboxes.append(False)
        
        #add destination column, the selected variable, and whether or not the column requires mapping to the "column_mapping_rules" list
        column_mapping_rules.append((item, var, child_dictionary['requires_mapping'],child_dictionary['is_percentage'])) #adds item, var, and requires mapping as as tupple
        
        row_num += 1
        
    print('Labels created')

    # Create a button to trigger the write_selections function
    button = tk.Button(field_map_window, text="Implement Field Mapping Rules", command=write_field_mappings)
    # button.grid(row=row_num, column=0, columnspan=2)
    button.grid(row=row_num, column=0)

    
    button = tk.Button(field_map_window, text="Open Value Mapping Rules Window", command=open_value_mappings_window)
    button.grid(row=row_num, column=3, columnspan=2)
    

MainNavWindow = tk.Tk()
MainNavWindow.title('Acquistions Analysis Tool')
MainNavWindow.geometry('300x300+50+50')

    
global bold_font
bold_font = tkfont.Font(weight='bold')

button = tk.Button(MainNavWindow, text='Clear TMT All', command=btn_temp)
button.grid(row=0, column=0)
button = tk.Button(MainNavWindow, text='Open Map Tape Fields Window', command=open_field_map_window)
button.grid(row=1, column=0)

# Start the Tkinter event loop
MainNavWindow.mainloop()
