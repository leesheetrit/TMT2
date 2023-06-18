#https://customtkinter.tomschimansky.com/documentation/
#https://github.com/tomschimansky/customtkinter


#approach:
    #get field mapping window loaded, no button functionality
    #move to seperate files
    #keep building
    #build functionality for base tool as needed

import customtkinter
import tkinter as tk
import openpyxl as xl
from tkinter import ttk

class NavigationFrame(customtkinter.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)

        # add widgets onto the frame, for example:
        self.label = customtkinter.CTkLabel(self, text="Navigation Window")
        self.label.grid(row=0, column=0, padx=20)
        self.button1 = customtkinter.CTkButton(self, text="Read Excel TMT File", \
                                               command=self.read_excel_TMT)
        self.button1.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.button2 = customtkinter.CTkButton(self, text="Map Tape Fields", \
                                               command=self.open_field_mapping_frame)
        self.button2.grid(row=2, column=0, padx=10, pady=5, sticky="ew")
        self.button3 = customtkinter.CTkButton(self, text="Map Tape Values")
        self.button3.grid(row=3, column=0, padx=10, pady=5, sticky="ew")
        
    def read_excel_TMT(self):
        print('Reading Excel TMT...')
        #call loading screen class
        print('Excel TMT Read')
        #call excel loaded class
        
        print('Loading Excel File...')
        self.wb_path = 'I:/ServicingMSR/QRM/Trades Upload Files/Tape Mapping Tool/Tape Mapping Tool.xlsx'
        self.wb = xl.load_workbook(self.wb_path)
        self.dest_columns_sheet = self.wb['Destination Columns']
        self.tape_sheet = self.wb['Tape']
        self.dest_field_values_sheet = self.wb['Destination Field Values']
        self.values_mapping_sheet = self.wb['ValuesMapping']
        #Find the last used row on the Excel sheet
        print('Finding Last Row on Tape File...')
        self.last_row = self.tape_sheet.max_row
        
        ####Read Source Columns 
        # Create a dictionary to store the source columns
        #stores the source column letter as the Dictionary item and the source Column Name as the dictionary value
        self.source_columns = {}
        for cell in self.tape_sheet[1]:
            if cell.value is not None:
                self.source_columns[cell.column_letter] = cell.value
                    
        print('Source Columns Read.')          
        # Sort the source column names alphabetically
        self.source_column_names = sorted(map(str, self.source_columns.values()))

        print('Source Columns Sorted.')          

        ####Read Destination Columns 
        #Read the destination columns into a dictionary called "destination_columns"
        #dictionarys are Key-Value pairs
        #stores the destination column name as the Dictionary Key
        self.destination_columns = {}
        for row in self.dest_columns_sheet.iter_rows(min_row=2, min_col=0, max_col=4):
            dest_col_name = row[0].value
            requires_mapping = row[1].value
            dest_value_example = row[2].value
            is_percentage = row[3].value
            
            #the value for the destination_columns dictionary is another dictionary
            #the key-value pairs of the child dictioary as shown below
            if dest_col_name is not None:
                self.destination_columns[dest_col_name] = {
                   'requires_mapping': requires_mapping,
                   'dest_value_example': dest_value_example,
                   'is_percentage': is_percentage,
                   }
        
    def open_field_mapping_frame(self):
        print('OpenFieldMappingFrame Function Activated')
        self.FieldMappingFrame= FieldMappingFrame(master=app)
        self.FieldMappingFrame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")

class FieldMappingFrame(customtkinter.CTkScrollableFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)

        # add widgets onto the frame, for example:
        self.label = customtkinter.CTkLabel(self, text = "Destination Column", \
                                            font=customtkinter.CTkFont(size=20, weight="bold"))
        self.label.grid(row=0, column=0, padx=20)

        self.label = customtkinter.CTkLabel(self, text = "Requires Mapping", \
                                            font=customtkinter.CTkFont(size=20, weight="bold"))
        self.label.grid(row=0, column=1, padx=20)
        
        self.label = customtkinter.CTkLabel(self, text = "One Loan Example", \
                                            font=customtkinter.CTkFont(size=20, weight="bold"))
        self.label.grid(row=0, column=2, padx=20)
               
        self.label = customtkinter.CTkLabel(self, text = "Source Column", \
                                            font=customtkinter.CTkFont(size=20, weight="bold"))
        self.label.grid(row=0, column=3, padx=20)
              
        self.label = customtkinter.CTkLabel(self, text = "Special", \
                                            font=customtkinter.CTkFont(size=20, weight="bold"))
        self.label.grid(row=0, column=4, padx=20)
        
        #temporarily showing this at the top
        # row_num = 1
        # button = tk.Button(field_map_window, text="Save Mapping Rules", command=open_save_load_field_mappings_window)    
        # # button.grid(row=row_num, column=0, columnspan=2)
        # button.grid(row=row_num, column=1)
        
        row_num = 2
        #create a "combos" list
        self.column_mapping_rules = []
        
        self.value_mapping_rules = []
        
        #### Gen. Widgets for Col. Mapping Window
        # Create a list to store the DivideBy100Var variables
        
        self.DivideBy100Vars = []
        self.AllCurrentVar = None 
        
        for item, child_dictionary in app.navigation_frame.destination_columns.items():
            # Create a label and combo box for each Destination Column in the "destination_columns" 
            self.label = customtkinter.CTkLabel(self, text=item)
            self.label.grid(row=row_num, column=0)
            
            if child_dictionary['requires_mapping'] == 'FALSE':
                RequiresMappingLabelText = ''
            else:
                RequiresMappingLabelText= child_dictionary['requires_mapping']
            #label2= tk.Label(self, text=child_dictionary['requires_mapping'])
            self.label2= customtkinter.CTkLabel(self, text=RequiresMappingLabelText)
            self.label2.grid(row=row_num, column=1)
            
            self.label3= customtkinter.CTkLabel(self, text=child_dictionary['dest_value_example'])
            self.label3.grid(row=row_num, column=2)
            
            #source columns combo box
            # var = tk.StringVar(self)
            self.combobox_var = customtkinter.StringVar()
            self.combo = customtkinter.CTkComboBox(self, variable=self.combobox_var, \
                                              values=list(app.navigation_frame.source_column_names), \
                                                  state ="readonly")
            self.combo.grid(row=row_num, column=3)
            
              
            # combo.bind('<KeyPress>', keyboard_first_letter_select)
               
            # Check if any of the source column names match "item" and set it as the default value
            #comment this out if you don't want the default matching
            for source_column_name in app.navigation_frame.source_column_names:
                if source_column_name.lower() == item.lower():
                    self.combobox_var.set(source_column_name)
                    break
            
            # check_var = customtkinter.StringVar(value="on")
            # checkbox = customtkinter.CTkCheckBox(app, text="CTkCheckBox", command=checkbox_event,
            #                          variable=check_var, onvalue="on", offvalue="off")
            
            
            if child_dictionary['is_percentage'] == True:
                # Create a variable to hold the state of the checkbox
                self.DivideBy100Var = tk.BooleanVar()
                # Create the checkbox
                self.checkbox = customtkinter.CTkCheckBox(self, text='Divide by 100?', variable=self.DivideBy100Var)
                self.checkbox.grid(row=row_num, column=4)
                self.DivideBy100Vars.append(self.DivideBy100Var) # Add the DivideBy100Var to the DivideBy100Vars list
            elif item == 'Delq':
                self.AllCurrentVar = tk.BooleanVar()
                checkbox = customtkinter.CTkCheckBox(self, text='All Current?', variable=self.AllCurrentVar)
                checkbox.grid(row=row_num, column=4)
                self.DivideBy100Vars.append(False)

            else:
                self.DivideBy100Vars.append(False)

            
            #add destination column, the selected variable, and whether or not the column requires mapping to the "column_mapping_rules" list
            self.column_mapping_rules.append((item, self.combobox_var, \
                                              child_dictionary['requires_mapping'],\
                                                  child_dictionary['is_percentage'])) #adds item, var, and requires mapping as as tupple
            
            row_num += 1
            
        print('Labels created')

        # # Create a button to trigger the write_selections function
        # # button = customtkinter.CTkButton(self, text="Implement Field Mapping Rules", command=write_field_mappings)
        # button = customtkinter.CTkButton(self, text="Implement Field Mapping Rules")
        # # button.grid(row=row_num, column=0, columnspan=2)
        # button.grid(row=row_num, column=0)

        
        # # button = customtkinter.CTkButton(self, text="Open Value Mapping Rules Window", command=open_value_mappings_window)
        # button = customtkinter.CTkButton(self, text="Open Value Mapping Rules Window")
        # button.grid(row=row_num, column=3, columnspan=2)
        
        ####Load Tape Mapping Tool Excel File
        
        
class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Tape Mapping Tool")
        self.geometry("700x700x5x5")
        self.grid_rowconfigure(0, weight=1)  # configure grid system
        self.grid_columnconfigure(1,weight = 1) 
        self.navigation_frame  = NavigationFrame(master=self)
        self.navigation_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        
      

if __name__ == "__main__":
    customtkinter.set_appearance_mode("dark")
    app = App()
    app.mainloop()


