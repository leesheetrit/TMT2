# -*- coding: utf-8 -*-
"""
Created on Thu Jun 29 07:50:31 2023

@author: lee.sheetrit
"""

import customtkinter
import tkinter as tk
import pandas as pd
from tkinter import ttk
import openpyxl as xl

class PrepTMT(customtkinter.CTkScrollableFrame):
    
    def __init__(self,master,wb_path, wb, tape_sheet,values_mapping_sheet,\
                 headers_merge_need = False, **kwargs):
        super().__init__(master, **kwargs)
        # super().__init__(master, **kwargs)
        
        print('Prep Excel TMT Class initiated')
        #Attributes
        self.wb_path = wb_path
        self.wb = wb
        self.tape_sheet = tape_sheet
        self.values_mapping_sheet = values_mapping_sheet

        self.raw_file_path = ''
        self.raw_wb = ''
        self.raw_wb_sheet_names = ''
        self.raw_wb_LL_sheet = ''
        self.headers_merge_need = headers_merge_need
        
        #Call Methods
        self.clear_file()
        self.select_LL_sheet()
        
    def clear_file(self):
        
        # Clear the "Tape" sheet
        # self.tape_sheet.delete_rows(1, self.tape_sheet.max_row)
        # if "Tape" in self.wb.sheetnames:
        # # Delete the "Tape" sheet
        if "Tape" in self.wb.sheetnames:
            self.wb.remove(self.wb["Tape"])
        
        # Clear rows 2 and onwards in the "ValuesMapping" sheet
        self.values_mapping_sheet.delete_rows(2, self.values_mapping_sheet.max_row)

        # Save the workbook
        self.wb.save(self.wb_path)
        print('Excel File Has been Reset')

    def select_LL_sheet(self):
        # Prompt the user for a file path for an excel file
        self.raw_file_path = tk.filedialog.askopenfilename(title='Select Excel File', filetypes=[('Excel Files', '*.xlsx')])
        
        print('Reading Excel Sheet Names...')
        if not self.raw_file_path:
            print('No file selected. Exiting.')
            return
        
        # Load the workbook
        self.raw_wb = xl.load_workbook(filename=self.raw_file_path)
        
        #print(self.raw_wb.sheetnames)            

        label = customtkinter.CTkLabel(self, text="Choose the Loan Level Sheet")
        label.grid(row=0, column=0)
        optionmenu = customtkinter.CTkOptionMenu(self, values=self.raw_wb.sheetnames,
                                         command=self.move_sheet)
        
        # sheet_listbox = Listbox(self)
        optionmenu.grid(row=1, column=0)


    def move_sheet(self, choice):
        print('moving sheet,',choice)
        #self.wb.copy_worksheet(self.raw_wb[choice])
        
        destination_sheet = self.wb.create_sheet('Tape')
        # destination_sheet.title = 'Tape'
        
        self.raw_wb_LL_sheet = self.raw_wb[choice]
        
        print('raw_wb[choice]:', self.raw_wb[choice])
        print('self.raw_wb_LL_sheet:', self.raw_wb_LL_sheet)
        
        for row in self.raw_wb_LL_sheet.iter_rows(values_only=True):
            # print(row)
            destination_sheet.append(row)

        if self.headers_merge_need:
            print('merging headers')
            #self.merge_headers()
            for column in range(1, destination_sheet.max_column + 1):
                
                # Get the values from row 1 and row 2
                value1 = destination_sheet.cell(row=1, column=column).value
                value2 = destination_sheet.cell(row=2, column=column).value
            # Concatenate the values in row 1 and row 2
                #cell_value = str(destination_sheet.cell(row=1, column=column).value) + ' ' + str(destination_sheet.cell(row=2, column=column).value)
                
                if value1 is not None:
                    cell_value = value1 + ' ' +  value2
                else:
                    cell_value =  value2
                # Update the cell value in row 1
                destination_sheet.cell(row=1, column=column).value = cell_value
            
            # Delete row 2
            destination_sheet.delete_rows(2)

        self.wb.save(self.wb_path)
        print("Success", "Sheet copied and renamed as 'tape'.")
        

     
