#https://customtkinter.tomschimansky.com/documentation/
#https://github.com/tomschimansky/customtkinter


#approach:
    #keep building
    #build functionality for base tool as needed

import customtkinter
import tkinter as tk
import openpyxl as xl
from tkinter import ttk

from PrepTMT import PrepTMT
from FieldMappingFrame import FieldMappingFrame
from ValueMappingFrame import ValueMappingFrame
from FieldMappingToDBFrame import FieldMappingToDBFrame
from sqlalchemy import create_engine
#import sql_library

class NavigationFrame(customtkinter.CTkFrame):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)

        widgets = []
        #Lavel
        widgets.append(customtkinter.CTkLabel(self, text="Navigation Window"))
        #Buttons
        
        #need to troubleshoot it seems that once I do this, the open field mappings window is still using the old tape file
        widgets.append(customtkinter.CTkButton(self, text="Prep Excel TMT File", \
                                               command=self.master.prep_TMT))
        
        widgets.append(customtkinter.CTkButton(self, text="Read Excel TMT File", \
                                               command=self.master.read_excel_TMT))    
        widgets.append(customtkinter.CTkButton(self, text="Open Field Mappings Window", \
                                               command=self.master.open_field_mapping_frame))
        widgets.append(customtkinter.CTkButton(self, text="Open Field Mappings Window w/ Smart Match", \
                                                   command=self.master.open_field_mapping_frame_smart))
        widgets.append(customtkinter.CTkButton(self, text="Write Field Mappings to xls", \
                                              command=self.master.write_field_mappings))
            
        widgets.append(customtkinter.CTkButton(self, text="Open Save Field Mappings Window", \
                                                  command=self.master.open_field_mappings_to_db_frame))
        widgets.append(customtkinter.CTkButton(self, text="Save Field Mappings to DB", \
                                                      command=self.master.save_field_mappings_to_db))  
            
        widgets.append(customtkinter.CTkButton(self, text="Open Value Mappings Window",\
                                               command=self.master.open_value_mapping_frame))
        widgets.append(customtkinter.CTkButton(self, text="Write Value Mappings", \
                                              command=self.master.write_value_mappings))
    
            
        
        for i, widget in enumerate(widgets):
            widget.grid(row=i, column=0, padx=10, pady=5, sticky="ew")

        
class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("Tape Mapping Tool")
        self.geometry("1600x700x5x5")
        self.grid_rowconfigure(0, weight=1)  # configure grid system
        self.grid_columnconfigure(1,weight = 1) 
        self.navigation_frame  = NavigationFrame(master=self)
        self.navigation_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        
        self.work_bench_frame = customtkinter.CTkFrame(master=self)
        self.work_bench_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
        
        # self.work_buttons_frame = customtkinter.CTkFrame(master=self)
        # self.work_buttons_frame.grid(row=2, column=1, padx=20, pady=20, sticky="nsew")
        
        self.source_columns = {}
        self.destination_columns = {}
        self.source_columns_sorted = []
        self.column_mappings_rules = []
        self.sellers_name = ''
        self.mappings_name = ''
        
        self.wb_path = 'I:/ServicingMSR/QRM/Trades Upload Files/Tape Mapping Tool/Tape Mapping Tool.xlsx'
        self.wb = xl.load_workbook(self.wb_path)
        self.dest_columns_sheet = self.wb['Destination Columns']
        self.tape_sheet = self.wb['Tape']
        self.dest_field_values_sheet = self.wb['Destination Field Values']
        self.values_mapping_sheet = self.wb['ValuesMapping']
        
        self.qrm_db= create_engine('mssql+pyodbc://CFAWSQRMSQL01/Servicing?driver=SQL+Server+Native+Client+11.0',
            fast_executemany = True)
        
    def prep_TMT(self):
        'Resetting Excel File'
        
        self.prepTMT_Window = PrepTMT(master=self,\
                            wb_path = self.wb_path,\
                            wb=self.wb,\
                            tape_sheet = self.tape_sheet,\
                            values_mapping_sheet = self.values_mapping_sheet)

        self.prepTMT_Window.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
    
    def read_excel_TMT(self):
        print('Reading Excel TMT...')
        #to be: call loading screen class
        print('Excel TMT Read')
        #to be: call excel loaded class
        

        #Find the last used row on the Excel sheet
        print('Finding Last Row on Tape File...')
        self.last_row = self.tape_sheet.max_row
        
        ####Read Source Columns 
        # Create a dictionary to store the source columns
        #stores the source column letter as the Dictionary item and the source Column Name as the dictionary value

        for cell in self.tape_sheet[1]:
            if cell.value is not None:
                self.source_columns[cell.column_letter] = cell.value
        print('Source Columns Read.')          
        # Sort the source column names alphabetically
        self.source_columns_sorted = sorted(map(str, self.source_columns.values()))
        # print(type(self.source_column_names))
        print('Source Columns Sorted.')          

        ####Read Destination Columns 
        #Read the destination columns into a dictionary called "destination_columns"
        #dictionarys are Key-Value pairs
        #stores the destination column name as the Dictionary Key
        for row in self.dest_columns_sheet.iter_rows(min_row=2, min_col=0, max_col=4):
            dest_col_name = row[0].value
            requires_mapping = row[1].value
            dest_value_example = row[2].value
            is_percentage = row[3].value
            
            #the value for the destination_columns dictionary is another dictionary
            #the key-value pairs of the child dictioary as shown below
            if dest_col_name is not None:
                self.destination_columns[dest_col_name] = {
                   'requires_mapping': requires_mapping,
                   'dest_value_example': dest_value_example,
                   'is_percentage': is_percentage,
                   }
        print('Excel TMT has been Loaded')
        
    def open_field_mapping_frame(self):
        print('OpenFieldMappingFrame Method Activated')
        self.work_bench_frame= FieldMappingFrame(master=self,\
                                                 destination_columns=self.destination_columns, \
                                                     source_columns =  self.source_columns_sorted )
        self.work_bench_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
         
        self.column_mappings_rules = self.work_bench_frame.get_field_mappings()
        
    def open_field_mapping_frame_smart(self):
        print('OpenFieldMappingFrame Method Activated')
        self.work_bench_frame= FieldMappingFrame(master=self,\
                                                 destination_columns=self.destination_columns, \
                                                     source_columns =  self.source_columns_sorted,
                                                     smart_match = True,\
                                                         qrm_db = self.qrm_db)
        self.work_bench_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
         
        self.column_mappings_rules = self.work_bench_frame.get_field_mappings()
        
    def write_field_mappings (self):
        ####Write Field mappings
        
            self.DivideBy100Vars = self.work_bench_frame.get_DivideBy100Vars()
            self.AllCurrentVar = self.work_bench_frame.get_AllCurrentVar()
            
            # Print the index and the values of the variables in DivideBy100Vars
            for idx, var in enumerate(self.DivideBy100Vars):
                # print(f"DivideBy100Vars: Index: {idx}, Value: {var}")
                if var is False:
                    print(f"DivideBy100Vars: Index: {idx}, Value: {var}")
                else:
                    print(f"DivideBy100Vars: Index: {idx}, Value: {var.get()}")      
            i=-1
            # Write the destination column names going across the columns of last_row
            for i, (item, var, requires_mapping,is_percentage, hard_code) in enumerate(self.column_mappings_rules):
              self.tape_sheet.cell(row=self.last_row+2, column=i+1).value = item
            print('Destination Column Labels writted into XLS')
            
            i=0
            row = self.last_row + 3
            # Skip to the next row and write the associated source column formulas or mappings across that row, one per column
            print('All Loans are Current equals:', self.AllCurrentVar.get())
            for i, (item, var, requires_mapping,is_percentage,hard_code) in enumerate(self.column_mappings_rules):
                col_letter = None
                for letter, name in self.source_columns.items():
                    if name == var.get():
                        col_letter = letter
                        break
                    
                print('Evaluating logic for',item,'All current: ',self.AllCurrentVar.get(), "i equals ", i)
                print('Evaluating logic for',item,'Hard Code: ',hard_code.get())
                
                # if self.DivideBy100Vars[i] is False:
                #     #no checkbox for this item
                #     print(f"Item: {item}, Index: {i}, DivideBy100Vars Checkbox Value: {self.DivideBy100Vars[i]}")
                # else:
                #     #read whether the checkbox was checked or not
                #     print(f"Item: {item},  Index: {i},DivideBy100Vars Checkbox Value: {self.DivideBy100Vars[i].get()}")
                     
                #if a source column is selected
                #or delinquent and all current
                #or there is hard code
                if col_letter is not None \
                    or (item =='Delq' and self.AllCurrentVar.get())\
                        or hard_code.get() != '':
                #if col_letter is not None or (item =='Delq' and self.AllCurrentVar.get()):
                    #if the checkbox for all curent is checked
                    if hard_code.get() != '':
                        formula =  hard_code.get()
                    elif item == "Delq" and self.AllCurrentVar.get():
                        #write all current
                        formula = "PREPAID OR CURRENT"
                    elif requires_mapping== 'TRUE' and col_letter is not None:
                        print('vlookup for ',item)
                        formula = f'=VLOOKUP({col_letter}$1'+ ' & "-" & ' +f'{col_letter}2 & "-" & "{item}", ValuesMapping!$A:$B, 2, FALSE)'       
                    #no divide by 100 Checkbox exists
                    elif self.DivideBy100Vars[i] is False:
                        formula = f"={col_letter}2"
                    else:
                        #if divide by 100 is checked
                        if self.DivideBy100Vars[i].get():
                            formula = f"={col_letter}2/100"
                        #divide by 100 is not checked
                        else:
                            formula = f"={col_letter}2"
                        
                    self.tape_sheet.cell(row=row, column=i+1).value = formula
            
            # Save the changes to the Excel file
            self.wb.save(self.wb_path)
            print('Mappings Written')
            

    def open_field_mappings_to_db_frame(self):
        self.field_mappings_to_db_frame= FieldMappingToDBFrame(master=self)
        self.field_mappings_to_db_frame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
    
    def save_field_mappings_to_db(self):
        
        #1) not quite working, only seems to be writing FIRST Due to First Due
        
        print('in method save_field_mappings_to_db')
        self.sellers_name = self.field_mappings_to_db_frame.get_seller_name()
        print (self.sellers_name)
        self.mappings_name = self.field_mappings_to_db_frame.get_mapping_name()
        print (self.mappings_name)
           
        for i, (dest_field, source_field, requires_mapping,is_percentage, hard_code) in enumerate(self.column_mappings_rules):
           
            if source_field.get() is not None and source_field.get() != "":
                insert_query = f"INSERT INTO Servicing.dbo.TMT_Field_Mappings \
                    (Sellers_Name, Mappings_Name, Destination_Field, Source_Field) \
                        VALUES \
                            ('{self.sellers_name}', '{self.mappings_name}', '{dest_field}', '{source_field.get()}')"
                self.qrm_db.connect().execute(insert_query)
                print('Wrote: ', dest_field,' - ', source_field.get())
        
        print('Field Mappings Written to Database')
    
    def open_value_mapping_frame(self):
        print('OpeValueMappingFrame Method Activated')
        self.ValueMappingFrame = ValueMappingFrame(master=self,
                                                last_row=self.last_row,\
                                                tape_sheet=self.tape_sheet,\
                                                column_mapping_rules=self.column_mappings_rules,
                                                dest_field_values_sheet= self.dest_field_values_sheet,\
                                                source_columns= self.source_columns,\
                                                wb_path = self.wb_path)
        self.ValueMappingFrame.grid(row=0, column=1, padx=20, pady=20, sticky="nsew")
    
    def write_value_mappings (self):
        ####Write Value mappings
        #get value_mapping_rules
        # pass
        self.value_mapping_rules = self.ValueMappingFrame.get_field_mappings()
        i=0
        for i, (source_col_letter,source_col_name,unique_source_value1,dest_field_name,dest_field_value) in enumerate(self.value_mapping_rules):
         
            self.values_mapping_sheet.cell(row=i+2, column=1).value = str(source_col_name) + '-' + str(unique_source_value1) + '-' + str(dest_field_name)
            print('writing')
            print(dest_field_name)
            self.values_mapping_sheet.cell(row=i+2, column=2).value = dest_field_value.get()
            
        self.wb.save(self.wb_path)
        print('Values mappings writted into excel')

        
        
if __name__ == "__main__":
    customtkinter.set_appearance_mode("dark")
    app = App()
    app.mainloop()


